# -*- coding: utf-8 -*-
from openpyxl import *
#import math

class Log:

    def lf(defect):
# <------------------ this loops is for the present and must be work with data log in future ! -----> MAYBE!!!
        if defect == '1':
            df = 'Пошкодження поверхні контакту'
        elif defect == '2':
            df = 'Гострини на розрізі контакту'
        elif defect == '3':
            df = 'Не симетричне / не відповідне закриття ядра'
        elif defect == '4':
            df = 'Пошкодження контакту або його деформація'
        elif defect == '5':
            df = 'Асиметрія контакту'
        elif defect == '6':
            df = 'Зазубрини в місті відрізу контакту'
        elif defect == '7':
            df = 'Пошкодження проводу або ущільнення'
        elif defect == '8':
            df = 'Рекваліфікація / EMPB'
        elif defect == '9':
            df = 'Не вірна довжина проводу'
        elif defect == '10':
            df = 'Не відповідне зварне з’єднання'
        elif defect == '11':
            df = 'Не відповідна сила стягування кабельбіндера'
        elif defect == '12':
            df = 'Не відповідне термоусадження'
        elif defect == '13':
            df = 'Не відповідне скручення проводів'
        elif defect == '14':
            df = 'Не відповідна обмотка з’єднання  '
        elif defect == '15':
            df = 'Тріснута запчастина'
        elif defect == '16':
            df = 'Обладнання не вмикається / не продукує виріб'
        elif defect == '17':
            df = 'інше'
        else:
            df = 'не вірно визначений дефект'
        # print(df)
        return df
# ------------------>

    def f_mark_row(ws):
        mark = '**'
        for row in ws:
            for cell in row:
                if cell.value == mark:
                    ex = cell.row
                    # print(ex)
                    return ex

    def mark(sap_eq1):
        wb = load_workbook('eq_log/eq_file/{}.xlsx'.format(sap_eq1))
        ws = wb['main']
        # find last mark row, for new data place
        ex = Log.f_mark_row(ws)
        return ex

    def te(type_eq):
        if type_eq == 'Аплікатор':
            type_eq_m = 'applicator'
        elif type_eq == 'Komax Alpha 355 / 355 S':
            type_eq_m = 'komax'
        elif type_eq == 'Komax Gamma 333 PC':
            type_eq_m = 'komax'
        elif type_eq == 'Schunk / Stapla ультразвукова зварка':
            type_eq_m = 'schunk'
        else:
            type_eq_m = 'other'
        return type_eq_m

    def markl(type_eq_m):
        w = load_workbook('eq_log/general equipment log.xlsx')
        sheet = str(type_eq_m)
        ws = w[sheet]
        # find last mark row, for new data place
        ex_log = Log.f_mark_row(ws)
        return ex_log

    def filtr(type_eq):
        if type_eq == 'Аплікатор':
            f = str('B25')
        elif type_eq == 'Komax Alpha 355 / 355 S':
            f = str('B26')
        elif type_eq == 'Komax Gamma 333 PC':
            f = str('B27')
        elif type_eq == 'Schunk / Stapla ультразвукова зварка':
            f = str('B28')
        elif type_eq == 'Кабельбіндеровий пістолет':
            f = str('H25')
        elif type_eq == 'Raychem / Raychem TE':
            f = str('H26')
        elif type_eq == 'Komax Twist BT 188 t / BT 188':
            f = str('H27')
        elif type_eq == 'Kabateck / Ondal':
            f = str('H28')
        else:
            f = str('B29')
        return f

    def type(sap_eq):
        sap_eq1 = '*8000' + str(sap_eq)
        w = load_workbook('list_eq.xlsx')
        ws = w['list_eq']
        # find last mark row, for new data place
        mark = sap_eq1
        for row in ws:
            for cell in row:
                if cell.value == mark:
                    ex_list = cell.row

        pos = 'C' + str(ex_list)
        type_eq = ws[pos].value
        return type_eq

    def cntr(sap_eq1, ex, counter):  # calculation of forecast repair
        wb = load_workbook('eq_log/eq_file/{}.xlsx'.format(sap_eq1))
        ws = wb['main']
        pos = str('D') + str(ex-1)       
        wrt_counter = ws[pos].value
        wrt_counter = int(wrt_counter)
        pos_aver = str('H2')
        num_aver = int(ws[pos_aver].value)
        counter = int(counter)
        aa = counter - wrt_counter
        b = num_aver - aa
        b = str(b)
        return b

    def l_cntr(sap_eq1, ex):  # find counter of last repair
        wb = load_workbook('eq_log/eq_file/{}.xlsx'.format(sap_eq1))
        ws = wb['main']
        pos = str('D') + str(ex-1)       
        a = ws[pos].value
        return a

    def eu_log(exl, type_eq_m, date, sap_eq1, per_number, df, counter):

        w = load_workbook('eq_log/general equipment log.xlsx')
        a = str('A') + str(exl)
        b = str('B') + str(exl)
        c = str('C') + str(exl)
        d = str('D') + str(exl)
        e = str('E') + str(exl)
        a1 = str('A') + str(exl + 1)  # for mark symbols
        # write data to equipment file
        sheet = w.get_sheet_by_name(type_eq_m)
        sheet[a] = date
        sheet[b] = sap_eq1
        sheet[c] = per_number
        sheet[d] = df
        sheet[e] = counter
        sheet[a1] = '**'
        w.save('eq_log/general equipment log.xlsx')

    def search_num_eq(sap_eq1, ex):
        wb = load_workbook('eq_log/eq_file/{}.xlsx'.format(sap_eq1))  # find file by number
        ws = wb['main']
        pos = str('D') + str(ex - 1)
        num = str(ws[pos].value)
        return num

    def data_log_eq(ex, wb, date, per_number, df, counter):
        a = str('A') + str(ex)
        b = str('B') + str(ex)
        c = str('C') + str(ex)
        d = str('D') + str(ex)
        ex1 = int(ex)
        a1 = str('A') + str(ex1 + 1)  # for mark symbols
        # write data to equipment file
        sheet = wb.get_sheet_by_name('main')
        sheet[a] = date
        sheet[b] = per_number
        sheet[c] = df
        sheet[d] = counter
        sheet[a1] = '**'
