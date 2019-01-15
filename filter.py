# -*- coding: utf-8 -*-
# write and read excel files with check the existence of the file

from openpyxl import *
import f_w_r

def shift_id(per_number):    # return name of shift by person number
    wb = load_workbook('files\personfile.xlsx')
    ws = wb['main']
    mark = str('*') + str(per_number)
    sheet = 'main'
    name = 'personfile'
    ex = f_w_r.find_row_fls(name, sheet, mark)
    if not ex:
        return False
    pos = str('D') + str(ex)
    sft = (ws[pos].value)
    return sft

def check_apl(sap_eq):    # check equipment is applicator
    if sap_eq and sap_eq.isdigit() and len(sap_eq) == 4:
        sap_eq1 = '*8000' + str(sap_eq)
        w = load_workbook('files\list_apl.xlsx')
        ws = w['main']
        for row in ws:
            for cell in row:
                if cell.value == sap_eq1:
                        return True
    return False

def write_apl_data(apl, date, per_number, type_fix, notice):
    if not apl.isdigit():
        return False
    apl = '8000' + str(apl)
    type_fix_massiv = ['механічне налаштування', 'заміна запчастин', 'налаштування симетричності розрізу']
    if type_fix in type_fix_massiv:
        f_w_r.write_eq_file(apl, date, per_number, type_fix, notice)

