# -*- coding: utf-8 -*-
import sys
# window1 is our interface file
from window2 import *
from openpyxl import *
from datetime import datetime


class MyWin(QtWidgets.QDialog):
    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.ui.pushButton.clicked.connect(self.action)  # print data - work
        self.ui.pushButton_2.clicked.connect(self.print_card)  # print data 1 page
        self.ui.pushButton_3.clicked.connect(self.print_card2)  # print data 2 page

    def action(self):
        self.ui.total_cycles.setText("")  # total_cycles
        self.ui.next_page.setText("")  # next page
        self.ui.correction.setText("")  # correction
        self.ui.total_cycles1.setText("")  # Sum of cycles in card

        # input data
        applicator = self.ui.applicator.toPlainText()
        cycles = self.ui.cycles.toPlainText()
        f_cycles = self.ui.f_cycles.toPlainText()
        initial_cycles = self.ui.initial_cycles.toPlainText()
        page = self.ui.page.toPlainText()
        counter = self.ui.counter.toPlainText()

        page1 = str(int(page) + 1)
        self.ui.next_page.setText(page1)

        total_cycles1 = str((int(cycles) + int(initial_cycles)) - int(f_cycles))
        self.ui.total_cycles1.setText(total_cycles1)

        counter = int(counter)
        if counter == 0:
            total_cycles = str(int(cycles) + int(initial_cycles))
            correction = str(0)
        else:
            total_cycles = str(counter)
            correction = str(int(counter) - (int(cycles) + int(initial_cycles)))

        # total_cycles = str(int(cycles) + int(initial_cycles) + int(counter))
        # correction = str(int(counter) - int(total_cycles))

        self.ui.total_cycles.setText(total_cycles)
        self.ui.correction.setText(correction)

        y_date = datetime.strftime(datetime.now(), "%Y")

        # write data to file
        wb = load_workbook(filename='counter_ver.1.xlsx', read_only=True)
        ws = wb[y_date]

        mark = "**"  # word to search
        for row in ws:
            for cell in row:
                if cell.value == mark:
                    x = cell.row  # find last mark row
                    # print(x)
                    break

        wrfile = load_workbook('counter_ver.1.xlsx')

        date = datetime.strftime(datetime.now(), "%Y/%m/%d")

        a = str('A') + str(x)
        b = str('B') + str(x)
        c = str('C') + str(x)
        d = str('D') + str(x)
        e = str('E') + str(x)
        a1 = str('A') + str(x + 1)

        sheet = wrfile.get_sheet_by_name(y_date)
        sheet[a] = int(applicator)
        sheet[b] = date
        sheet[c] = int(page)
        sheet[d] = int(cycles)
        sheet[e] = int(total_cycles)
        sheet[a1] = '**'

        wrfile.save('counter_ver.1.xlsx')

        # write data to template file, 1 page
        wrfile = load_workbook('Template_apl_cycles.xlsx')

        b = str('B') + str(8)
        i = str('I') + str(8)
        f = str('F') + str(44)

        sheet = wrfile.get_sheet_by_name('print1')
        sheet[b] = y_date
        sheet[i] = int(applicator)
        sheet[f] = int(page1)

        wrfile.save('Template_apl_cycles.xlsx')

        # write data to template file, 2 page
        wrfile = load_workbook('Template_apl_cycles2.xlsx')

        d = str('D') + str(30)
        b = str('B') + str(35)

        sheet = wrfile.get_sheet_by_name('print2')
        sheet[b] = "correction is: {}".format(correction)
        sheet[d] = int(total_cycles) + int(correction)

        wrfile.save('Template_apl_cycles2.xlsx')

    # print 2 sheets in 2 file
    def print_card(self):
        os.startfile("Template_apl_cycles.xlsx", "print")  # write correct address, file name

    def print_card2(self):
        os.startfile("Template_apl_cycles2.xlsx", "print")  # write correct address, file name


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    myapp = MyWin()
    myapp.show()
sys.exit(app.exec_())
