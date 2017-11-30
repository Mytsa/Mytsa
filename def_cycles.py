# -*- coding: utf-8 -*-
from openpyxl import *
from datetime import datetime

applicator = int(input('input SAP number of applicator: '))
date = datetime.strftime(datetime.now(), "%Y/%m/%d")
page = int(input('input page number of card: '))
cycles = int(input('input cycles: '))
old_cycles = int(input('input cycles previous card: '))
c = input('has a counter (y/no): ')

def counter(cycles, old_cycles):
    if c is 'y':  # another if-else
        counter = int(input('input counter data: '))
        correction = counter - (cycles + old_cycles)
        # print('correction is: {} '.format(correction))
        return correction
    else:
        total_cycles = cycles + old_cycles
        return total_cycles

# output block
def output(total_cycles, old_cycles, cycles, page, correction):
    print("total cycles is {}".format(total_cycles))
    print("sum of cycles in this cart is {}".format(old_cycles + cycles))
    # print('date is {}'.format(date))
    print('next page is: ', page + 1)
    if c == 'y':
        print('correction is: ', correction)

def exc():
    wb = load_workbook(filename='counter_ver.1.xlsx', read_only=True)
    ws = wb['2017']

    mark = "**"  # word to search
    for row in ws:
        for cell in row:
            if cell.value == mark:
                x = cell.row  # find last mark row
                return x

def wrt_exc(y, x, arg):
    wrfile = load_workbook('counter_ver.1.xlsx')
    sheet = wrfile.get_sheet_by_name('2017')
    y = str('Y') + str(x)
    arg = sheet[y]
    return arg


# a = str('A') + str(x)
# b = str('B') + str(x)
# c = str('C') + str(x)
# d = str('D') + str(x)
# e = str('E') + str(x)
# a1 = str('A') + str(x + 1)
#
# sheet = wrfile.get_sheet_by_name('2017')
# sheet[a] = applicator
# sheet[b] = date
# sheet[c] = page
# sheet[d] = cycles
# sheet[e] = total_cycles
# sheet[a1] = '**'

wrfile.save('counter_ver.1.xlsx')

# print file
# import os
#
# os.startfile("D:/1.txt", "print")  # write corect adress and name of file
# #

#
# class Data():
#     def __init__(self):
#         self.applicator = int(input('input SAP number of applicator: '))
#         self.date = datetime.strftime(datetime.now(), "%Y/%m/%d")
#         self.page = int(input('input page number of card: '))
#         self.cycles = int(input('input cycles: '))
#         self.old_cycles = int(input('input cycles previous card: '))
#         self.c = input('has a counter (y/no): ')
#
#     def io(self):
#         applicator = self.applicator
#         date = self.date
#         page = self.page
#         cycles = self.cycles
#         old_cycles = self.old_cycles
#         c = self.c
#
#
# class Exc():
#     def __init__(self):
#         self.wb = load_workbook(filename='counter_ver.1.xlsx', read_only=True)
#         self.ws = wb['2017']
#
#     def open(self):
#         wb = self.wb
#         ws = self.ws
#
#     def mark(self):
#         mark = "**"  # word to search
#         for row in ws:
#             for cell in row:
#                 if cell.value == mark:
#                     x = cell.row  # find last mark row
#                     break
#


