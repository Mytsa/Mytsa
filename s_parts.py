# -*- coding: utf-8 -*-
from openpyxl import *
from datetime import datetime

class Parts:
    # def __init__(self):
    #     self.w_date = datetime.strftime(datetime.now(), "%W")

    def f_mark_row(ws):
        mark = '**'
        for row in ws:
            for cell in row:
                if cell.value == mark:
                    ex = cell.row
                    return ex

    def f_mark_row_weld(ws):    # search row for welding spare parts
        mark = '***'
        for row in ws:
            for cell in row:
                if cell.value == mark:
                    ex1 = cell.row
                    return ex1

    def mark(w_date, per_number):
        wb = load_workbook('spare_parts/{}.xlsx'.format(per_number))
        ws = wb[w_date]
        ex = Parts.f_mark_row(ws)
        return ex

    def wrt_templ(per_number, w_date, px, num_part, name_part, pcs, sap_eq1, counter):
        wrfile1 = load_workbook('spare_parts/{}.xlsx'.format(per_number))
        sheet = wrfile1.get_sheet_by_name(w_date)
        # index coordinate for template file
        a = str('B') + str(px)
        b = str('C') + str(px)
        c = str('D') + str(px)
        d = str('E') + str(px)
        e = str('F') + str(px)  # counter
        a1 = str('B') + str(1+px)
        # write data to template file
        sheet[a] = num_part
        sheet[b] = name_part
        sheet[c] = pcs
        sheet[d] = int(sap_eq1)
        sheet[e] = int(counter)
        sheet[a1] = '**'
        wrfile1.save(('spare_parts/{}.xlsx'.format(per_number)))

    def mark_log(a):
        wb = load_workbook('spare_parts/s_parts_log.xlsx')
        ws = wb[a]
        # find last mark row, for new data place
        ex = Parts.f_mark_row(ws)
        return ex

    def wrt_log(px, date, per_number, num_part, name_part, pcs, sap_eq1, counter):
        wrfile2 = load_workbook('spare_parts/s_parts_log.xlsx')
        sheet = wrfile2.get_sheet_by_name('main')
        px = int(px)    # only for position A1
        # index coordinate for template file
        a = str('A') + str(px)
        b = str('B') + str(px)
        c = str('C') + str(px)
        d = str('D') + str(px)
        e = str('E') + str(px)
        f = str('F') + str(px)
        g = str('G') + str(px)
        a1 = str('A') + str(1+px)
        # write data to template file
        sheet[a] = date
        sheet[b] = per_number
        sheet[c] = num_part
        sheet[d] = name_part
        sheet[e] = pcs
        sheet[f] = sap_eq1
        sheet[g] = counter
        sheet[a1] = '**'
        wrfile2.save('spare_parts/s_parts_log.xlsx')

    def apl_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number):
        pcs = int(pcs)    # for calculation in excel
        wb = load_workbook('spare_parts/spare_parts_per_month.xlsx')
        sheet = wb.get_sheet_by_name(m_date)
        ws = wb[m_date]
        px = Parts.f_mark_row(ws)
        px = int(px)    # only for applicators spare parts position A1
        # index coordinate for file
        a = str('A') + str(px)
        b = str('B') + str(px)
        c = str(name_part) + str(px)
        e = str('E') + str(px)
        f = str('F') + str(px)
        g = str('G') + str(px)
        a1 = str('A') + str(1+px)
        # write data to file
        sheet[a] = sap_eq1
        sheet[b] = date
        sheet[c] = pcs
        sheet[e] = counter
        sheet[f] = df
        sheet[g] = per_number
        sheet[a1] = '**'  # ** for applicators spare parts , *** - for schunk spare parts

        sheet = wb.get_sheet_by_name('summary')
        sheet["B18"] = date
        wb.save('spare_parts/spare_parts_per_month.xlsx')

    def weld_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number):
        pcs = int(pcs)  # for calculation in excel
        wb = load_workbook('spare_parts/spare_parts_per_month.xlsx')
        sheet = wb.get_sheet_by_name(m_date)
        ws = wb[m_date]
        px1 = Parts.f_mark_row_weld(ws)
        px1 = int(px1)  # only for applicators spare parts position A1
        # index coordinate for file
        a = str('A') + str(px1)
        b = str('B') + str(px1)
        c = str(name_part) + str(px1)
        g = str('g') + str(px1)
        h = str('h') + str(px1)
        o = str('i') + str(px1)
        a1 = str('A') + str(1 + px1)
        # write data to file
        sheet[a] = sap_eq1
        sheet[b] = date
        sheet[c] = pcs
        sheet[g] = counter
        sheet[h] = df
        sheet[o] = per_number
        sheet[a1] = '***'  #*** - for schunk spare parts

        sheet = wb.get_sheet_by_name('summary')
        sheet["B18"] = date
        wb.save('spare_parts/spare_parts_per_month.xlsx')


    def name_part(mark):
        wb = load_workbook('spare_parts/spare_parts_catalog.xlsx')
        ws = wb['main']
        mark = str('*') + str(mark)
        for row in ws:
            for cell in row:
                if cell.value == mark:
                    ex = cell.row
                    pos = 'C' + str(ex)
                    name = ws[pos].value
                    return name

    def position_part(mark):
        wb = load_workbook('spare_parts/spare_parts_catalog.xlsx')
        ws = wb['main']
        mark = str('*') + str(mark)
        for row in ws:
            for cell in row:
                if cell.value == mark:
                    ex = cell.row
                    pos = 'D' + str(ex)
                    position = ws[pos].value
                    return position

    def sap_part(mark):
        wb = load_workbook('spare_parts/spare_parts_catalog.xlsx')
        ws = wb['main']
        mark = str('*') + str(mark)
        for row in ws:
            for cell in row:
                if cell.value == mark:
                    ex = cell.row
                    pos = 'A' + str(ex)
                    position = ws[pos].value
                    return position

    def check_name(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number):
        if name_part == ('crimper wire'):  # write data to file with control of repair of applicator
            name_part = 'C'
            Parts.apl_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number)
        elif name_part == ('crimper insulation'):
            name_part = 'C'
            Parts.apl_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number)

        elif name_part == ('anvil'):
            name_part = 'D'
            Parts.apl_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number)
        elif name_part == ('anvil wire'):
            name_part = 'D'
            Parts.apl_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number)
        elif name_part == ('anvil insulation'):
            name_part = 'D'
            Parts.apl_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number)

        elif name_part == ('anvil plate'):
            name_part = 'C'
            Parts.weld_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number)
        elif name_part == ('sonotrode'):
            name_part = 'D'
            Parts.weld_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number)
        elif name_part == ('gilding jaw'):
            name_part = 'E'
            Parts.weld_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number)
        elif name_part == ('anvil weld'):
            name_part = 'F'
            Parts.weld_sp(m_date, sap_eq1, date, pcs, name_part, df, counter, per_number)

        else:
            pass
