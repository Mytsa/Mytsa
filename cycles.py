# -*- coding: utf-8 -*-
import xlrd  # read excel file
import xlwt  # write excel file
import xlsxwriter
from datetime import datetime

# applicator = int(input('input SAP number of applicator: '))
# date = datetime.strftime(datetime.now(), "%Y/%m/%d ")
# page = int(input('input page number of card: '))
# cycles = int(input('input cycles: '))
# old_cycles = int(input('input cycles previous card: '))
# c = input('has a counter (y/no): ')
#
# if c is 'y':
#     counter = int(input('input counter data: '))
#     correction = counter - (cycles + old_cycles)
#     total_cycles = counter
#     # print('correction is: {} '.format(correction))
# else:
#     total_cycles = cycles + old_cycles

# wb = xlwt.Workbook()    # initialization new file to write
# ws = wb.add_sheet('2017')    # initialization new file with sheet
#
# #open file
# rb = xlrd.open_workbook('mark.xls', formatting_info=True)
# # choice active sheet
# sheet = rb.sheet_by_index(0)
# i = 0
# while sheet.row_values(i)[0] != '*':  # find last row before mark --> **
#     x = i + 1
#     i += 1
# print(x)
#
# ws.write(x + 1, 0, "*")  # mark last row with symbol --> **
# wb.save('mark.xls')

# --------------------------- new openpyxl
import openpyxl
from openpyxl import *

wb = load_workbook(filename='counter_ver.1.xlsx', read_only=True)
ws = wb['mark']

for row in ws.rows:
    for cell in row:
        a = cell.value
print(a)
i = 1
j = 1
while a != '*':
    for row in ws.rows:
        for cell in row:
            x = i
            i += 1
            print(i)

###            find cell location
#for i in range(1,ws.max_row):
#    if ws.cell(row=row, column=0).value == date:
#        for j in range(i, ws.max_column):
#            print (ws.cell(row=i, column=j).value)



#rom openpyxl import load_workbook  #https://stackoverflow.com/questions/10939450/openpyxl-basic-search
#import glob
#import re
#import os
#wb = load_workbook('xl.xlsx')
#sh = wb.get_sheet_by_name('Russian')
#for name in glob.glob('*.mp3'):
#    audio = EasyID3(name)
#    for row_index in range(sh.get_highest_row()):
#       if sh.cell(row=row_index, column=0).value == name:
#            print(row_index)


#openpyxl.worksheet.Worksheet.iter_cols()
#for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
#    for cell in row:
#        print(cell)

#openpyxl.worksheet.Worksheet.rows()
#>>> ws = wb.active
#>>> ws['C9'] = 'hello world'
#>>> tuple(ws.rows)



   ###         
# xfile = openpyxl.load_workbook('counter_ver.1.xlsx')
#
# x = 2
#
# a = str('A') + str(x)
# b = str('B') + str(x)
# c = str('C') + str(x)
# d = str('D') + str(x)
# e = str('E') + str(x)
#
# sheet = xfile.get_sheet_by_name('2017')
# sheet[a] = applicator
# sheet[b] = date
# sheet[c] = page
# sheet[d] = cycles
# sheet[e] = total_cycles
#
# xfile.save('counter_ver.1.xlsx')
