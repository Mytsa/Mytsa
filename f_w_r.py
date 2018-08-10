# -*- coding: utf-8 -*-
# write and read excel files with check the existence of the file

from openpyxl import *

def open_file(name, sheet):
    wb = load_workbook('files\eq_files\{}.xlsx'.format(name))
    ws = wb[sheet]
    return ws
#<!--- if you need know about file at open moment

    # try:
    #     wb = load_workbook('files/{}.xlsx'.format(name))
    #     ws = wb[sheet]
    #     return ws
    # except FileNotFoundError:
    #     mes = ('file not find')
#-------------!>

def find_row(name, sheet, mark):
    try:
        # open_file(name, sheet):
        for row in open_file(name, sheet):
            for cell in row:
                if cell.value == mark:
                    ex = cell.row
                    # print(ex)
                    mes = 'file is find'
                    #print(mes)
                    return ex
    except FileNotFoundError:    # if you work with filter in def open_file - use this error "FileNotFoundError:"
        mes = 'file not find, please create file or check you data input'
        #print(mes)
        return mes

#<!-------- block for check open file action

# mark = 'Gamma 1 (8_1235)'
# name = 'log'
# sheet = 'main'
# find_row(name, sheet, mark)

#------------------!>


def find_row_fls(name, sheet, mark):
    try:
        wb = load_workbook('files\{}.xlsx'.format(name))
        ws = wb[sheet]
        for row in ws:
            for cell in row:
                if cell.value == mark:
                    ex = cell.row
                    return ex
    except FileNotFoundError:    # if you work with filter in def open_file - use this error "FileNotFoundError:"
        mes = 'file not find, please create file or check you data input'
        return mes


def find_data_in_row(sheet, name, mark, type_fix):

    ex = find_row_fls(name, sheet, mark)

    wb = load_workbook('files\{}.xlsx'.format(name))
    ws = wb[sheet]
    pos = str(type_fix) + str(ex)
    num = str(ws[pos].value)
    return num


def clean_sum_by_weeks(sheet):
    w = load_workbook('files\DownTime_shift.xlsx')
    # ws = wb['main']
    # sheet = 'main'
    sheet = w.get_sheet_by_name(sheet)


    # <! ------ WRITE sum of figures to komax table
    letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    for i in range(3, 18):
        for j in letters:
            b = str(j) + str(i)
            write = '0'
            sheet[b] = int(write)

    # <! ------ WRITE sum of figures to schunk table
    letters = ['T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA']
    for i in range(3, 19):
        for j in letters:
            b = str(j) + str(i)
            write = '0'
            sheet[b] = int(write)

    w.save('files\DownTime_shift.xlsx')


def write_eq_file(eq_number, date, per_number, type_fix, notice):
    sheet = 's_p'
    mark = '**'
    ex = find_row(eq_number, sheet, mark)
    w = load_workbook('files\eq_files\{}.xlsx'.format(eq_number))

    a = str('A') + str(ex)
    b = str('B') + str(ex)
    c = str('C') + str(ex)
    d = str('D') + str(ex)
    a1 = str('A') + str(ex + 1)  # for mark symbols
    # write data to equipment file
    sheet = w.get_sheet_by_name('s_p')
    sheet[a] = date
    sheet[b] = per_number
    sheet[c] = type_fix
    sheet[d] = notice
    sheet[a1] = '**'    # mark last row for find in new write process

    w.save('files\eq_files\{}.xlsx'.format(eq_number))

def write_sum_and_shift_files(eq_number, type_fix, minutes):    # need sum of figures in table for creat summary table
    name = 'DownTime_shift'
    sheet = 'main'
    mark = '*' + str(eq_number)

    ex = find_row_fls(name, sheet, mark)
    pos = pos_by_fix(type_fix)
    data_from_pos = find_data_in_row(sheet, name, mark, pos)
    f_data = int(data_from_pos) + int(minutes)
    b = str(pos) + str(ex)

    # write data to equipment file
    w = load_workbook('files\{}.xlsx'.format(name))
    sheet = w.get_sheet_by_name(sheet)
    sheet[b] = f_data
    w.save('files\{}.xlsx'.format(name))


def write_sum_by_weeks(w_date):
    wb = load_workbook('files\DownTime_shift.xlsx')
    ws = wb['main']
    w = load_workbook('files\summary.xlsx')
    sheet = w.get_sheet_by_name(str(w_date))

# <! ------ WRITE sum of figures to komax table
    letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    for i in range(3, 18):
        for j in letters:
            pos = str(j) + str(i)
            figures = str(ws[pos].value)
            data_from_pos = str(sheet[pos].value)
            f_sum = int(figures) + int(data_from_pos)
            b = str(j) + str(i)
            sheet[b] = f_sum

# <! ------ WRITE sum of figures to schunk table
    letters = ['T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA']
    for i in range(3, 19):
        for j in letters:
            pos = str(j) + str(i)
            figures = str(ws[pos].value)
            data_from_pos = str(sheet[pos].value)
            f_sum = int(figures) + int(data_from_pos)
            b = str(j) + str(i)
            sheet[b] = f_sum

    w.save('files\summary.xlsx')

def write_log_file(date, shift, eq_number, apl, minutes, data, notice):    # write data to Log file
    name = 'log'
    sheet = 'main'
    mark = '**'

    ex = find_row_fls(name, sheet, mark)
    w = load_workbook('files\{}.xlsx'.format(name))

    a = str('A') + str(ex)
    b = str('B') + str(ex)
    c = str('C') + str(ex)
    d = str('D') + str(ex)
    e = str('E') + str(ex)
    f = str('F') + str(ex)
    g = str('G') + str(ex)
    a1 = str('A') + str(ex + 1)  # for mark symbols

    # write data to equipment file
    sheet = w.get_sheet_by_name(sheet)
    sheet[a] = date
    sheet[b] = shift
    sheet[c] = int(eq_number)
    sheet[d] = apl
    sheet[e] = int(minutes)
    sheet[f] = data
    sheet[g] = notice
    sheet[a1] = '**'    # mark last row for find in new write process

    w.save('files\{}.xlsx'.format(name))


def write_type(type):
    file = open('files\write_type.txt', 'w')
    file.write(type)
    file.close()


def data_from_write_type():
    f = open('files\write_type.txt', 'r')
    type_fix = f.read()
    return type_fix



def pos_by_fix(type_fix):

    # komax of DownTime

    if type_fix == 'проблеми з матеріалом':
        pos = 'B'
        return pos
    elif type_fix == 'механічна поломка':
        pos = 'C'
        return pos
    elif type_fix == 'електрична поломка':
        pos = 'D'
        return pos
    elif type_fix == 'очікування тех.відділу':
        pos = 'E'
        return pos
    elif type_fix == 'ПЗ':
        pos = 'F'
        return pos
    elif type_fix == 'СPU 2000':
        pos = 'G'
        return pos
    elif type_fix == 'scaner':
        pos = 'H'
        return pos
    elif type_fix == 'ТО обладнання':
        pos = 'I'
        return pos
    elif type_fix == 'інший тип простою':
        pos = 'J'
        return pos

    # applicators type of DownTime

    elif type_fix == 'механічне налаштування':
        pos = 'K'
        return pos
    elif type_fix == 'заміна запчастин':
        pos = 'L'
        return pos
    elif type_fix == 'налаштування симетричності розрізу':
        pos = 'M'
        return pos
    elif type_fix == 'ТО аплікатора':
        pos = 'N'
        return pos

    # seal DownTime
    elif type_fix == 'налаштування втулочного модуля':
        pos = 'O'
        return pos

    # printer DownTime
    elif type_fix == 'налаштування принтера':
        pos = 'P'
        return pos

    # schunk type of DownTime

    elif type_fix == 'проблеми з матеріалом на зварці':
        pos = 'T'
        return pos
    elif type_fix == 'Збій програми':
        pos = 'U'
        return pos
    elif type_fix == 'новий проект/нові параметри':
        pos = 'V'
        return pos
    elif type_fix == 'заміна електродів':
        pos = 'W'
        return pos
    elif type_fix == 'чистка обладнання':
        pos = 'X'
        return pos
    elif type_fix == 'ремонт електродів / зазор':
        pos = 'Y'
        return pos
    elif type_fix == 'ТО зварки':
        pos = 'Z'
        return pos
    elif type_fix == 'очікування зварка':
        pos = 'AA'
        return pos
    else:
        pass