# -*- coding: utf-8 -*-
import sys
# window1 is our interface file
from window2 import *
from openpyxl import *
from datetime import datetime


class MyWin(QtWidgets.QDialog):
    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.ui.pushButton.clicked.connect(self.action)  # print data - work
        self.ui.pushButton_2.clicked.connect(self.print_card)  # print data 1 page
        self.ui.pushButton_3.clicked.connect(self.print_card2)  # print data 2 page


    def action(self):
        self.ui.total_cycles.setText("")  # total_cycles
        self.ui.next_page.setText("")  # next page
        self.ui.correction.setText("")  # correction
        self.ui.total_cycles1.setText("")  # Sum of cycles in card
        self.ui.sum_cycles.setText("")  # Sum of cycles in old_card

        # input data
        applicator = self.ui.applicator.toPlainText()
        cycles = self.ui.cycles.toPlainText()
        f_cycles = self.ui.f_cycles.toPlainText()
        initial_cycles = self.ui.initial_cycles.toPlainText()
        page = self.ui.page.toPlainText()
        counter = self.ui.counter.toPlainText()

        # calcuclation
        page1 = str(int(page) + 1)
        self.ui.next_page.setText(page1)

        
        total_cycles1 = str((int(cycles) + int(initial_cycles)) - int(f_cycles))
        sum_old = str((int(cycles)- int(f_cycles)))                   

        #total_cycles1 = str((int(cycles) + int(initial_cycles)) - int(f_cycles))
        self.ui.total_cycles1.setText(total_cycles1)
        self.ui.sum_cycles.setText(sum_old)

        # counter if loop
        counter = int(counter)
        if counter == 0:
            total_cycles = str(total_cycles1)
            correction = str(0)
        else:
            total_cycles = str(counter)
            correction = str(int(counter) - (int(total_cycles1)))

        #kof = float((int(total_cycles1)) / 30000)  # calculation of remainder in 30 000 cycles
        #first_row = (kof - int(kof)) * 30000    # cycles to write in first row - not good calculation with tolerance (+/- 5 000)


     
        # <--- need check data for input total cycles is more than last recorded in file !!!

        
        # end check data --->

        # write data after counter if loop
        self.ui.total_cycles.setText(total_cycles)
        self.ui.correction.setText(correction)

        # read/write data intro/to file

        y_date = datetime.strftime(datetime.now(), "%Y")    # local variables for excel
        date = datetime.strftime(datetime.now(), "%Y/%m/%d")    # local variables for excel

        wb = load_workbook(filename='counter_apl.xlsx', read_only=True)
        ws = wb[y_date]

        mark = "**"  # word to search last raw in sheet
        for row in ws:
            for cell in row:
                if cell.value == mark:
                    x = cell.row  # find last mark row, for data input place
                    break

        # write data to counter file
        wrfile = load_workbook('counter_apl.xlsx')

        # index coordinate
        a = str('A') + str(x)
        b = str('B') + str(x)
        c = str('C') + str(x)
        d = str('D') + str(x)
        e = str('E') + str(x)
        a1 = str('A') + str(x + 1)    # for mark symbols

        # write data
        sheet = wrfile.get_sheet_by_name(y_date)
        sheet[a] = int(applicator)
        sheet[b] = date
        sheet[c] = int(page)
        sheet[d] = int(cycles)
        sheet[e] = int(total_cycles)
        sheet[a1] = '**'

        wrfile.save('counter_apl.xlsx')

        # write/save data to template file, 1 page
        wrfile = load_workbook('Template_apl_cycles.xlsx')

        b = str('B') + str(8)
        i = str('I') + str(8)
        #e1 = str('E') + str(11)
        f = str('F') + str(44)

        sheet = wrfile.get_sheet_by_name('print1')
        sheet[b] = y_date
        sheet[i] = int(applicator)
        #sheet[e1] = int(first_row)
        sheet[f] = int(page1)

        wrfile.save('Template_apl_cycles.xlsx')

        # write/save data to template file, 2 page
        wrfile = load_workbook('Template_apl_cycles2.xlsx')

        d = str('D') + str(30)
        b = str('B') + str(35)

        sheet = wrfile.get_sheet_by_name('print2')
        sheet[b] = "correction is: {}".format(correction)
        sheet[d] = int(total_cycles)

        wrfile.save('Template_apl_cycles2.xlsx')


    # print 2 sheets in 2 file
    def print_card(self):
        os.startfile("Template_apl_cycles.xlsx", "print")  # write correct address, file name

    def print_card2(self):
        os.startfile("Template_apl_cycles2.xlsx", "print")  # write correct address, file name


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    myapp = MyWin()
    myapp.show()
sys.exit(app.exec_())
