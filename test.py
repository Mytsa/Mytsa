# from openpyxl import load_workbook
# wb2 = load_workbook('files/eq_files/80001841.xlsx')
# print(wb2.get_sheet_names('s_p'))

from openpyxl import *
from f_w_r import *

# def shift_id(mark):    # return name of shift by person number
#     # mark = '3070'
#     wb = load_workbook('files\personfile.xlsx')
#     ws = wb['main']
#     for row in ws:
#         for cell in row:
#             if cell.value == mark:
#                 ex = cell.row
#                 pos = ex
#                 print(pos)
#
# a = '*3070'
# print(shift_id(a))

per_number = '*3070'

# def shift_id(per_number):    # return name of shift by person number
#     name = 'personfile'   # file with data about person in shift
#     sheet = 'main'
#     mark = '*3070'
#     sft = find_data_in_row(name, sheet, mark)
#     print(sft)
#
#     return sft
# shift_id(per_number)


# wb = load_workbook('files\personfile.xlsx')
# ws = wb['main']
# mark = '*3070'
# sheet = 'main'
# name = 'personfile'
#
# ex = find_row_fls(name, sheet, mark)
# pos = str('D') + str(ex)
#
# print(ws[pos].value)
eq_num = input('imnput number : ')

a = eq_num.isdigit()
print(a)
# num = str(num)
# if eq_num == int(eq_num):
#     print('write correct equipment number')
#
# else:
#     print('good')